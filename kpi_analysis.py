import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tqdm import tqdm
import math
import os
from utils import *
from openpyxl import load_workbook
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import time

def write_to_excel_with_overwrite(df, file_path, sheet_name):
    # 使用pd.ExcelWriter保存DataFrame，指定openpyxl为引擎
    if os.path.isfile(file_path):
        # 使用openpyxl加载现有Excel文件
        book = load_workbook(file_path)
        
        # 如果工作表存在，则删除
        if sheet_name in book.sheetnames:
            del book[sheet_name]
        
        # 保存并关闭书籍
        book.save(file_path)
        book.close()
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    else:
        df.to_excel(file_path,sheet_name=sheet_name,index=False,header=False)   
        
        
def process_whole_life_plan(file_path,folder_path):
    df = pd.read_excel(file_path)
    grouped = df.groupby("本次维修时间")["工作包人天"].sum().reset_index()
    # 计算工作包人天的最大值
    max_value = grouped["工作包人天"].max()

    # 计算工作包人天的最小值
    min_value = grouped["工作包人天"].min()

    # 计算工作包人天的平均值
    mean_value = grouped["工作包人天"].mean()
    
    kpi_data = []
    fluctuation = []
    for index, row in grouped.iterrows():
        # 获取 "本次维修时间" 列的值
        repair_time = row["本次维修时间"]
        
        # 获取 "工作包人天" 列的值
        work_days = row["工作包人天"]
        kpi_data.append([repair_time, work_days/mean_value -1])
        fluctuation.append(abs(work_days/mean_value -1))
    kpi_data = [['季度工时波动上限','季度工时波动下限','季度工时波动平均值'],[max_value/mean_value-1,1-min_value/mean_value,sum(fluctuation)/len(fluctuation)],
                ['季度','季度工时波动']] + kpi_data
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    df = pd.DataFrame(kpi_data)
    write_to_excel_with_overwrite(df, os.path.join(folder_path, '季度-月度-日工时KPI统计.xlsx'), '季度工时')
        

def process_year_plan(file_path,folder_path, config):
    current_day = convert_str_to_date(config['today'])
    month_list = []
    cnt = 0
    current_month = convert_day_to_month(current_day)
    month_list.append(current_month)
    while cnt < 12-1:
        cnt += 1
        # 增加一个月
        if current_day.month == 12:
            current_day = current_day.replace(year=current_day.year + 1, month=1)
        else:
            current_day = current_day.replace(month=current_day.month + 1)
        
        current_month = convert_day_to_month(current_day)
        month_list.append(current_month)
    
    df = pd.read_excel(file_path)
    df = df[df['本次维修时间'].isin(month_list)]
    grouped = df.groupby("本次维修时间")["工作包人天"].sum().reset_index()
    # 计算工作包人天的最大值
    max_value = grouped["工作包人天"].max()

    # 计算工作包人天的最小值
    min_value = grouped["工作包人天"].min()

    # 计算工作包人天的平均值
    mean_value = grouped["工作包人天"].mean()
    
    kpi_data = []
    fluctuation = []
    for index, row in grouped.iterrows():
        # 获取 "本次维修时间" 列的值
        repair_time = row["本次维修时间"]
        
        # 获取 "工作包人天" 列的值
        work_days = row["工作包人天"]
        kpi_data.append([repair_time, work_days/mean_value -1])
        fluctuation.append(abs(work_days/mean_value -1))
    kpi_data = [['月度工时波动上限','月度工时波动下限','月度工时波动平均值'],[max_value/mean_value-1,1-min_value/mean_value,sum(fluctuation)/len(fluctuation)],
                ['月度','月度工时波动']] + kpi_data
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    df = pd.DataFrame(kpi_data)
    write_to_excel_with_overwrite(df, os.path.join(folder_path, '季度-月度-日工时KPI统计.xlsx'), '月度工时')     

def process_month_plan(file_path,folder_path, config):
    df = pd.read_excel(file_path)
    df = df[df['本次维修时间'].str[:7] == config['today'][:7]]
    
    grouped = df.groupby("本次维修时间")["工作包人天"].sum().reset_index()
    # 计算工作包人天的最大值
    max_value = grouped["工作包人天"].max()

    # 计算工作包人天的最小值
    min_value = grouped["工作包人天"].min()

    # 计算工作包人天的平均值
    mean_value = grouped["工作包人天"].mean()
    
    kpi_data = []
    fluctuation = []
    for index, row in grouped.iterrows():
        # 获取 "本次维修时间" 列的值
        repair_time = row["本次维修时间"]
        
        # 获取 "工作包人天" 列的值
        work_days = row["工作包人天"]
        kpi_data.append([repair_time, work_days/mean_value -1])
        fluctuation.append(abs(work_days/mean_value -1))
    kpi_data = [['日工时波动上限','日工时波动下限','日工时波动平均值'],[max_value/mean_value-1,1-min_value/mean_value,sum(fluctuation)/len(fluctuation)],
                ['日','日工时波动']] + kpi_data
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    df = pd.DataFrame(kpi_data)
    write_to_excel_with_overwrite(df, os.path.join(folder_path, '季度-月度-日工时KPI统计.xlsx'), '日工时')    

def process_interval(file_path,folder_path, config):
    df = pd.read_excel(file_path)
    df = df[['工作包间隔转换值', '本次维修时间', '上次维修时间']]
    df['本次维修时间'] = pd.to_datetime(df['本次维修时间'])
    df['上次维修时间'] = pd.to_datetime(df['上次维修时间'])

    # 计算两个日期的差值，并转换为天数
    df['维修间隔天数'] = (df['本次维修时间'] - df['上次维修时间']).dt.days
    df['工作包间隔转换值'] = pd.to_numeric(df['工作包间隔转换值'], errors='coerce')
    # 计算偏离修天数
    df['偏离修天数'] = df['工作包间隔转换值'] - df['维修间隔天数']
    df['偏离修百分比'] = df['偏离修天数'] / df['工作包间隔转换值']
    
    
    
    average_deviation_percentage = df['偏离修百分比'].abs().mean()
    average_interval_utilization = 1 - average_deviation_percentage
    
    # 计算过修百分比的平均值
    average_over_repair_percentage = df[df['偏离修百分比'] > 0]['偏离修百分比'].mean()

    # 计算欠修百分比的平均值
    average_under_repair_percentage = df[df['偏离修百分比'] < 0]['偏离修百分比'].mean()
    
    kpi_data = [['平均偏离修百分比','平均间隔利用率','平均过修百分比','平均欠修百分比'],
                [average_deviation_percentage,average_interval_utilization,average_over_repair_percentage,-average_under_repair_percentage]]
    
    # # 假设 df 是您的 DataFrame

    # # 按照维修天数分组
    # grouped = df.groupby('维修天数')

    # # 计算每组的平均偏离修百分比
    # average_deviation_percentage = grouped['偏离修百分比'].mean()

    # # 计算每组的平均间隔利用率
    # average_interval_utilization = 1 - average_deviation_percentage

    # # 计算每组的平均过修百分比（只考虑偏离修百分比大于0的情况）
    # average_over_repair_percentage = grouped.apply(lambda x: x[x['偏离修百分比'] > 0]['偏离修百分比'].mean())

    # # 计算每组的平均欠修百分比（只考虑偏离修百分比小于0的情况）
    # average_under_repair_percentage = grouped.apply(lambda x: x[x['偏离修百分比'] < 0]['偏离修百分比'].mean())

    # average_deviation_percentage_list = average_deviation_percentage.tolist()
    # average_interval_utilization_list = average_interval_utilization.tolist()
    # average_over_repair_percentage_list = average_over_repair_percentage.tolist()
    # average_under_repair_percentage_list = average_under_repair_percentage.tolist()

    # combined_results = list(zip(average_deviation_percentage_list, 
    #                         average_interval_utilization_list, 
    #                         average_over_repair_percentage_list, 
    #                         average_under_repair_percentage_list))

    
    
    # grouped = df.groupby("本次维修时间")["工作包人天"].sum().reset_index()
    # # 计算工作包人天的最大值
    # max_value = grouped["工作包人天"].max()

    # # 计算工作包人天的最小值
    # min_value = grouped["工作包人天"].min()

    # # 计算工作包人天的平均值
    # mean_value = grouped["工作包人天"].mean()
    
    # kpi_data = []
    # fluctuation = []
    # for index, row in grouped.iterrows():
    #     # 获取 "本次维修时间" 列的值
    #     repair_time = row["本次维修时间"]
        
    #     # 获取 "工作包人天" 列的值
    #     work_days = row["工作包人天"]
    #     kpi_data.append([repair_time, work_days/mean_value -1])
    #     fluctuation.append(abs(work_days/mean_value -1))
    # kpi_data = [['日工时波动上限','日工时波动下限','日工时波动平均值'],[max_value/mean_value-1,1-min_value/mean_value,sum(fluctuation)/len(fluctuation)],
    #             ['日','日工时波动']] + kpi_data
    # if not os.path.exists(folder_path):
    #     os.makedirs(folder_path)
    df = pd.DataFrame(kpi_data)
    
    write_to_excel_with_overwrite(df, os.path.join(folder_path, '季度-月度-日工时KPI统计.xlsx'), '间隔利用率')    



if __name__ == '__main__':
    start_time = time.time()
    config_name = './datasets/config.yaml'
    config = read_file_config(config_name)
    
    file_path = './results/全寿命计划.xlsx'
    folder_path = './results/kpi/'
    process_whole_life_plan(file_path,folder_path)
    file_path = './results/年计划.xlsx'
    process_year_plan(file_path,folder_path,config)
    file_path = './results/月计划.xlsx'
    process_month_plan(file_path,folder_path,config)
    process_interval(file_path,folder_path,config)
    
    
    end_time = time.time()
    elapsed_time_in_seconds = end_time - start_time
    hours, remainder = divmod(elapsed_time_in_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)

    # 打印结果
    print(f"程序运行时间：{int(hours)}时{int(minutes)}分{int(seconds)}秒")