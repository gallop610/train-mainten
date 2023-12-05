import openpyxl
import os
import multiprocessing as mp
import math
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from tqdm import tqdm
import pandas as pd
from utils.utils_time import convert_day_to_quarter, convert_day_to_month, quarter_difference, month_difference, add_quarters, add_months
        
def output_wholeLife_plan(workpackage):
    mainten_quarter = []
    for i in range(len(workpackage)):
        work = workpackage[i]
        work.Track_Type_Priority = sorted(work.Track_Type_Priority, key=lambda x: x[1], reverse=False)
        first_elements = [item[0] for item in work.Track_Type_Priority]
        track_type = ','.join(first_elements)
        for j in range(len(work.mainten_quarter)):
            date = work.mainten_quarter[j]
            if j == 0:
                last_mainten = convert_day_to_quarter(work.last_mainten_time)
            else:
                last_mainten = work.mainten_quarter[j-1]
                
            interval_quarter = quarter_difference(date,last_mainten)
            over_under_repaired = int(work.Work_Package_Interval_Conversion_Value/90)-quarter_difference(date,last_mainten)
            over_under_repaired_percentage = over_under_repaired/(int(work.Work_Package_Interval_Conversion_Value/90))
            # 红线过修
            red_line_over_date = add_quarters(last_mainten,int(math.ceil(work.Work_Package_Interval_Conversion_Value*0.90/90)))
            # 红线欠修
            red_line_repair_date = add_quarters(last_mainten,int(math.ceil(work.Work_Package_Interval_Conversion_Value*1.05/90)))
            tmp = [work.Train_Type,work.Train_Number,work.Work_Package_Number,work.Work_Package_Name,work.Electrifiable,work.Powerless,work.Work_Package_Interval_Time_Dimension,
                   work.Work_Package_Interval_Value,work.Work_Package_Interval_Conversion_Value,work.Work_Package_Person_Day,work.Work_Package_Person_Day_Unit_Price,work.Need_Trial_Run,
                   work.Work_On_Non_Working_Day,work.Cooling_Time,work.Shared_Cooling_Work_Package_Number,work.Sampling_Times,work.Combined_Work_Package_Number,work.Need_High_Voltage_Verification,
                   work.Work_Package_Contract,work.Undercarriage_Assembly,date,last_mainten,interval_quarter,over_under_repaired,over_under_repaired_percentage,red_line_over_date,red_line_repair_date,track_type]
            mainten_quarter.append(tmp)
            
    
    mainten_quarter.sort(key=lambda x: (x[20], x[8], x[1], x[2],x[21],x[13]), reverse=False)
    columns_name = ['列车类型','列车编号','工作包编号','工作包名称','是否可以通电工作','是否可以断电工作','工作包间隔时间维度','工作包间隔值','工作包间隔转换值', '工作包人天', '工作包人天单价',
               '是否需要试车','是否在工作日工作','冷却时间(天)','共享冷却工作包编号','抽检次数','组合工作包编号','是否需要高压验证','工作包承包','是否车下拆装','本次维修时间','上次维修时间',
               '间隔季度差值','过修季度','过修百分比','红线过修日期','红线欠修日期','股道类型']
    
    df = pd.DataFrame(columns=columns_name, data=mainten_quarter)
    folder_path = './results'
    # 判断文件夹是否存在，如果不存在则创建
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    df.to_excel(os.path.join(folder_path, '全寿命计划.xlsx'),sheet_name='全寿命计划',index=False)
    
    # wb = openpyxl.Workbook()
    # ws = wb.create_sheet(title='全寿命计划')
    # ws.append(['列车类型','列车编号','工作包编号','工作包名称','是否可以通电工作','是否可以断电工作','工作包间隔时间维度','工作包间隔值','工作包间隔转换值', '工作包人天', '工作包人天单价',
    #            '是否需要试车','是否在工作日工作','冷却时间(天)','共享冷却工作包编号','抽检次数','组合工作包编号','是否需要高压验证','工作包承包','是否车下拆装','本次维修时间','上次维修时间',
    #            '间隔季度差值','过修季度','过修百分比','红线欠修日期','红线过修日期','股道类型'])
    # mainten_quarter.sort(key=lambda x: (x[20], x[8], x[1], x[2],x[21],x[13]), reverse=False)
    # for data in mainten_quarter:
    #     ws.append(list(data))
    # wb.remove(wb['Sheet'])
    # folder_path = './results'
    # # 判断文件夹是否存在，如果不存在则创建
    # if not os.path.exists(folder_path):
    #     os.makedirs(folder_path)
    # wb.save(os.path.join(folder_path, '全寿命计划.xlsx'))
    
def output_year_plan(workpackage):
    mainten_month = []
    for i in range(len(workpackage)):
        work = workpackage[i]
        work.Track_Type_Priority = sorted(work.Track_Type_Priority, key=lambda x: x[1], reverse=False)
        first_elements = [item[0] for item in work.Track_Type_Priority]
        track_type = ','.join(first_elements)
        for j in range(len(work.mainten_month)):
            date = work.mainten_month[j]
            if j == 0:
                last_mainten = convert_day_to_month(work.last_mainten_time)
            else:
                last_mainten = work.mainten_month[j-1]

            interval_quarter = month_difference(date,last_mainten)
            over_under_repaired = int(work.Work_Package_Interval_Conversion_Value/30)-month_difference(date,last_mainten)- int(int(work.Work_Package_Interval_Conversion_Value)/2190)                                                                                                       
            over_under_repaired_percentage = over_under_repaired/(int(work.Work_Package_Interval_Conversion_Value/30))
            # 红线过修
            if work.Work_Package_Interval_Conversion_Value>90:
                red_line_over_date = add_months(last_mainten,int(math.ceil(work.Work_Package_Interval_Conversion_Value*0.90/90)))
            else:
                red_line_over_date = add_months(last_mainten,int(math.ceil(work.Work_Package_Interval_Conversion_Value*0.95/30)))
            # 红线欠修
            red_line_repair_date = add_months(last_mainten,int(math.ceil(work.Work_Package_Interval_Conversion_Value*1.05/90)))             
                
            tmp = [work.Train_Type,work.Train_Number,work.Work_Package_Number,work.Work_Package_Name,work.Electrifiable,work.Powerless,work.Work_Package_Interval_Time_Dimension,
                   work.Work_Package_Interval_Value,work.Work_Package_Interval_Conversion_Value,work.Work_Package_Person_Day,work.Work_Package_Person_Day_Unit_Price,work.Need_Trial_Run,
                   work.Work_On_Non_Working_Day,work.Cooling_Time,work.Shared_Cooling_Work_Package_Number,work.Sampling_Times,work.Combined_Work_Package_Number,work.Need_High_Voltage_Verification,
                   work.Work_Package_Contract,work.Undercarriage_Assembly,date,last_mainten,interval_quarter,over_under_repaired,over_under_repaired_percentage,red_line_over_date,red_line_repair_date,track_type]
            mainten_month.append(tmp)
    mainten_month.sort(key=lambda x: (x[20], x[8], x[1], x[2],x[21],x[13]), reverse=False)
    columns_name = ['列车类型','列车编号','工作包编号','工作包名称','是否可以通电工作','是否可以断电工作','工作包间隔时间维度','工作包间隔值','工作包间隔转换值', '工作包人天', '工作包人天单价',
               '是否需要试车','是否在工作日工作','冷却时间(天)','共享冷却工作包编号','抽检次数','组合工作包编号','是否需要高压验证','工作包承包','是否车下拆装','本次维修时间','上次维修时间',
               '间隔月份差值','过修月份','过修百分比','红线过修日期','红线欠修日期','股道类型']
    
    df = pd.DataFrame(columns=columns_name, data=mainten_month)
    folder_path = './results'
    # 判断文件夹是否存在，如果不存在则创建
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    df.to_excel(os.path.join(folder_path, '年计划.xlsx'),sheet_name='年计划',index=False)
    
    
    # wb = openpyxl.Workbook()
    # ws = wb.create_sheet(title='年计划')
    # ws.append(['列车类型','列车编号','工作包编号','工作包名称','是否可以通电工作','是否可以断电工作','工作包间隔时间维度','工作包间隔值','工作包间隔转换值', '工作包人天', '工作包人天单价',
    #            '是否需要试车','是否在工作日工作','冷却时间(天)','共享冷却工作包编号','抽检次数','组合工作包编号','是否需要高压验证','工作包承包','是否车下拆装','本次维修时间','上次维修时间',
    #            '间隔月份差值','过修月份','过修百分比','红线欠修日期','红线过修日期','股道类型'])
    # mainten_month.sort(key=lambda x: (x[20], x[8], x[1], x[2],x[21],x[13]), reverse=False)
    # for data in mainten_month:
    #     ws.append(list(data))
    # wb.remove(wb['Sheet'])
    # folder_path = './results'
    # # 判断文件夹是否存在，如果不存在则创建
    # if not os.path.exists(folder_path):
    #     os.makedirs(folder_path)
    # wb.save(os.path.join(folder_path, '年计划.xlsx'))
    
def output_month_plan(workpackage):
    mainten_day = []
    for i in range(len(workpackage)):
        work = workpackage[i]
        work.Track_Type_Priority = sorted(work.Track_Type_Priority, key=lambda x: x[1], reverse=False)
        first_elements = [item[0] for item in work.Track_Type_Priority]
        track_type = ','.join(first_elements)
        for j in range(len(work.mainten_day)):
            date = work.mainten_day[j]
            if j == 0:
                last_mainten = work.last_mainten_time
            else:
                last_mainten = work.mainten_day[j-1]
                
            interval_quarter = (date-last_mainten).days
            over_under_repaired = int(work.Work_Package_Interval_Conversion_Value)-(date-last_mainten).days
            over_under_repaired_percentage = over_under_repaired/int(work.Work_Package_Interval_Conversion_Value)
            # 红线过修
            if work.Work_Package_Interval_Conversion_Value>90:
                red_line_over_date = last_mainten + relativedelta(days=int(work.Work_Package_Interval_Conversion_Value*0.90))
            else:
                red_line_over_date = last_mainten + relativedelta(days=int(work.Work_Package_Interval_Conversion_Value*0.95))
            # 红线欠修
            red_line_repair_date = last_mainten + relativedelta(days=int(work.Work_Package_Interval_Conversion_Value*1.05))          
                
            tmp = [work.Train_Type,work.Train_Number,work.Work_Package_Number,work.Work_Package_Name,work.Electrifiable,work.Powerless,work.Work_Package_Interval_Time_Dimension,
                   work.Work_Package_Interval_Value,work.Work_Package_Interval_Conversion_Value,work.Work_Package_Person_Day,work.Work_Package_Person_Day_Unit_Price,work.Need_Trial_Run,
                   work.Work_On_Non_Working_Day,work.Cooling_Time,work.Shared_Cooling_Work_Package_Number,work.Sampling_Times,work.Combined_Work_Package_Number,work.Need_High_Voltage_Verification,
                   work.Work_Package_Contract,work.Undercarriage_Assembly,date.strftime("%Y-%m-%d"),last_mainten.strftime("%Y-%m-%d"),interval_quarter,over_under_repaired,over_under_repaired_percentage,red_line_over_date.strftime("%Y-%m-%d"),red_line_repair_date.strftime("%Y-%m-%d"),track_type]    
            mainten_day.append(tmp)
    mainten_day.sort(key=lambda x: (x[20], x[8], x[1], x[2],x[21],x[13]), reverse=False)        
    columns_name = ['列车类型','列车编号','工作包编号','工作包名称','是否可以通电工作','是否可以断电工作','工作包间隔时间维度','工作包间隔值','工作包间隔转换值', '工作包人天', '工作包人天单价',
               '是否需要试车','是否在工作日工作','冷却时间(天)','共享冷却工作包编号','抽检次数','组合工作包编号','是否需要高压验证','工作包承包','是否车下拆装','本次维修时间','上次维修时间',
               '间隔天数差值','过修天数','过修百分比','红线过修日期','红线欠修日期','股道类型']
    
    col = -1
    for index, info in enumerate(mainten_day):
        # if info[20] == '2024-04-01':
        if info[20] == '2026-12-31':
            col = index
            break
    mainten_day = mainten_day[:col]
    
    df = pd.DataFrame(columns=columns_name, data=mainten_day)
    folder_path = './results'
    # 判断文件夹是否存在，如果不存在则创建
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    df.to_excel(os.path.join(folder_path, '月计划.xlsx'),sheet_name='月计划',index=False)



    # wb = openpyxl.Workbook()
    # ws = wb.create_sheet(title='月计划')
    # ws.append(['列车类型','列车编号','工作包编号','工作包名称','是否可以通电工作','是否可以断电工作','工作包间隔时间维度','工作包间隔值','工作包间隔转换值', '工作包人天', '工作包人天单价',
    #            '是否需要试车','是否在工作日工作','冷却时间(天)','共享冷却工作包编号','抽检次数','组合工作包编号','是否需要高压验证','工作包承包','是否车下拆装','本次维修时间','上次维修时间',
    #            '间隔天数差值','过修天数','过修百分比','红线欠修日期','红线过修日期','股道类型'])
    # mainten_day.sort(key=lambda x: (x[20], x[8], x[1], x[2],x[21],x[13]), reverse=False)
    # print('正在写入月计划...')
    # for data in tqdm(mainten_day):
    #     if data[20] == '2024-04-01':
    #         break
    #     ws.append(list(data))
    # wb.remove(wb['Sheet'])
    # folder_path = './results'
    # # 判断文件夹是否存在，如果不存在则创建
    # if not os.path.exists(folder_path):
    #     os.makedirs(folder_path)
    # wb.save(os.path.join(folder_path, '月计划.xlsx'))